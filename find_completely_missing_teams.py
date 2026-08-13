"""
Find teams with COMPLETELY missing data for the 2025-2026 season.

A team is "completely missing" when BOTH are true:
  1. No game-by-game data  -> team is in data_gaps `teamsNoBoxScores` (or `errors`)
  2. No stat-section data   -> no team_total record with GP > 0 in the all_stats_tab file

Universe of teams comes from each state's *_data_gaps_<gender>_2025_2026.json.
Output: completely_missing_teams_25_26.csv  +  .xlsx
Columns: Team Name | Team URL | State | Season | Gender
"""
import json
import csv
import os

SEASON = "2025-2026"

# (Display name, folder name, file code)
STATES = [
    ("Arkansas",   "Arkansas_scraped_data",   "ar"),
    ("Indiana",    "Indiana_scraped_data",    "in"),
    ("Louisiana",  "Louisiana_scraped_data",  "la"),
    ("Michigan",   "Michigan_scraped_data",   "mi"),
    ("New Mexico", "NewMaxico_scraped_data",  "nm"),
    ("Ohio",       "Ohio_scraped_data",       "oh"),
    ("Oklahoma",   "Oklahoma_scraped_data",   "ok"),
    ("Texas",      "Texas_scraped_data",      "tx"),
    ("Washington", "Washington_scraped_data", "wa"),
]

GENDERS = [
    ("boys",  "Boys Basketball"),
    ("girls", "Girls Basketball"),
]

BASE = os.path.dirname(os.path.abspath(__file__))


def load(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def team_id_from_url(url):
    """https://www.maxpreps.com/wa/bickleton/bickleton-pirates/basketball/ -> wa/bickleton/bickleton-pirates/basketball"""
    if not url:
        return ""
    u = url.strip()
    for pref in ("https://www.maxpreps.com/", "http://www.maxpreps.com/", "https://maxpreps.com/"):
        if u.startswith(pref):
            u = u[len(pref):]
            break
    return u.strip("/")


def teams_with_real_stats(stats_path):
    """Set of team_ids that have a team_total record with GP > 0 in the stats tab."""
    have = set()
    if not os.path.exists(stats_path):
        return have
    data = load(stats_path)
    records = data if isinstance(data, list) else data.get("data", [])
    for rec in records:
        if not isinstance(rec, dict):
            continue
        if rec.get("record_type") != "team_total":
            continue
        try:
            gp = float(rec.get("GP", 0) or 0)
        except (TypeError, ValueError):
            gp = 0
        if gp > 0:
            tid = (rec.get("team_id") or "").strip("/")
            if tid:
                have.add(tid)
    return have


def main():
    rows = []
    summary = []

    for disp, folder, code in STATES:
        for gcode, glabel in GENDERS:
            gaps_path = os.path.join(BASE, folder, f"{code}_data_gaps_{gcode}_{SEASON.replace('-', '_')}.json")
            stats_path = os.path.join(BASE, folder, f"{code}_all_stats_tab_{gcode}_{SEASON.replace('-', '_')}.json")

            if not os.path.exists(gaps_path):
                summary.append(f"  [skip] {disp} {glabel}: no data_gaps file")
                continue

            gaps = load(gaps_path)
            has_stats = teams_with_real_stats(stats_path)

            # Candidates = no game-by-game data (no box scores) + teams that errored out
            candidates = list(gaps.get("teamsNoBoxScores", []))
            candidates += list(gaps.get("errors", []))

            missing_here = 0
            seen = set()
            for t in candidates:
                url = t.get("teamUrl", "")
                tid = team_id_from_url(url)
                if tid in seen:
                    continue
                seen.add(tid)
                # Exclude if it actually has stat-section stats
                if tid in has_stats:
                    continue
                rows.append({
                    "Team Name": t.get("teamName", "").strip(),
                    "Team URL": url.strip(),
                    "State": disp,
                    "Season": SEASON,
                    "Gender": glabel,
                })
                missing_here += 1

            summary.append(
                f"  {disp:12s} {glabel:16s}: {missing_here:4d} completely missing "
                f"(noBox={len(gaps.get('teamsNoBoxScores', []))}, errors={len(gaps.get('errors', []))}, "
                f"universe={gaps.get('meta', {}).get('totalTeams', '?')})"
            )

    # Sort: State, Gender, Team Name
    rows.sort(key=lambda r: (r["State"], r["Gender"], r["Team Name"].lower()))

    fields = ["Team Name", "Team URL", "State", "Season", "Gender"]

    csv_path = os.path.join(BASE, "completely_missing_teams_25_26.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    xlsx_path = os.path.join(BASE, "completely_missing_teams_25_26.xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Missing Teams 25-26"
        ws.append(fields)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r[k] for k in fields])
        widths = {"A": 38, "B": 70, "C": 14, "D": 12, "E": 18}
        for col, wdt in widths.items():
            ws.column_dimensions[col].width = wdt
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:E{len(rows) + 1}"
        wb.save(xlsx_path)
        xlsx_ok = True
    except Exception as e:
        xlsx_ok = False
        xlsx_err = str(e)

    print("=" * 70)
    print("Per state/gender breakdown:")
    print("\n".join(summary))
    print("=" * 70)
    print(f"TOTAL completely-missing teams: {len(rows)}")
    print(f"CSV : {csv_path}")
    if xlsx_ok:
        print(f"XLSX: {xlsx_path}")
    else:
        print(f"XLSX: FAILED ({xlsx_err})")


if __name__ == "__main__":
    main()
